import serial  # install with " pip3 install pyserial"
import time
import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
from PySide6.QtWidgets import QFileDialog, QMessageBox


def file_writter(serial_port1, serial_port2, ser1, ser2, output_file): 
    data1 = None
    data2 = None

        # Open the file to write the data
    
    try:
        while True:
            #with file_lock:
                if ser1.in_waiting > 0 : 
                    # Read a line from the serial port
                    data1 = ser1.readline().decode('utf-8').strip() 
                    print(f"Received from {serial_port1}: {data1}")
                    ##timestamp = datetime.datetime.now()
                    # Print the line to the console
                    #myDataLine =timestamp.strftime("%d-%m-%Y %H:%M:%S")+","+ line1 
                
                    # Write the line to the file
                    # file.write(myDataLine + '\n')

                if ser2.in_waiting > 0:
                    data2 = ser2.readline().decode('utf-8').strip()
                    print(f"received from{serial_port2}: {data2}")

                if data1 is not None and data2 is not None:
                    timestamp = datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S")
                    if(len(data1.split(","))) >= 3 :
                        myDataLine = f"{timestamp},{data1},{data2}"
                        print(f"Writing to file: {myDataLine}")
                        with open(output_file, 'a') as file:
                            file.write(myDataLine + '\n')
                            file.flush()  # Ensure the data is written to the file immediately
                            file.close()
                        # Reset the data after writing
                        data1 = None
                        data2 = None
                        time.sleep(1)  # Add a small delay to prevent high CPU usage
                #     else:
                #         pass  
                
                # if data1 is not None and data2 is None

                else:
                    print("No data waiting in the serial buffer.")
                time.sleep(1)  # Add a small delay to prevent high CPU usage
    except KeyboardInterrupt:
        print("Program interrupted. Closing...")
    except serial.SerialException as e:
        print(f"Serial communication error: {e}")
    finally:
       # print(f"Disconnected from {serial_port1} and {serial_port2}.")   
       if ser1 :
           QMessageBox.warning(None , "Warning", "Micron 1", buttons=QMessageBox.Ok, defaultButton=QMessageBox.NoButton)



def data_lesen(lese_datei):
        # Öffnen der Datei im Lesemodus
        try:
            with open(lese_datei, "r") as myfile:
                # Initialisieren einer leeren Liste, um die Zeilen zu speichern
                lines = []
                
                # Durchlaufen jeder Zeile in der Datei
                for line in myfile:
                    # Entfernen des Zeilenumbruchs und Hinzufügen zur Liste
                    lines.append(line.strip())
                
                # Rückgabe der Liste mit den Zeilen
                return lines
        except FileNotFoundError:
            print(f"Datei {lese_datei} nicht gefunden.")
            return []
        

def data_lesen_zeitraum(lese_datei,date1, date2, dialog):
    options = QFileDialog.Options()
    file_path, _ = QFileDialog.getSaveFileName(dialog, "Speichern unter", "", "Excel Dateien (*.xlsx);;Alle Dateien (*)", options=options)
    date1_formated = date1.date().toPython()
    date2_formated = date2.date().toPython()
   
    
    if file_path:
        try:
            data_fromTo = []  
            myData = data_lesen(lese_datei)  # Assuming data_lesen is a method in your class
            for line in myData: 
                line_splited = line.split(",")
                if len(line_splited)>=6:
                    line_date = datetime.datetime.strptime(line_splited[0].split(" ")[0], "%d.%m.%Y").date()
                    if date1_formated <= line_date <= date2_formated:
                        data_fromTo.append(line_splited)
                        

            wb = Workbook()
            ws = wb.active
            today = datetime.datetime.now().strftime("%d.%m.%Y")
            title_document = ws.cell(row=3,column=6, value=(f"BATGUARD DATA: {today}"))

            header_color = PatternFill(start_color="1E90FF", end_color="1E90FF", fill_type="solid")  # DodgerBlue3 color
            cell_color = PatternFill(start_color="98F5FF", end_color="98F5FF", fill_type="solid")
            
            thin_border = Border(left=Side(style='thin', color='000000'),
                                 right=Side(style='thin', color='000000'),
                                 top=Side(style='thin', color='000000'),
                                 bottom=Side(style='thin', color='000000'))
            #print("data_splited", data_fromTo)

            headers = ['Datum','Uhrzeit','Einfluege', 'Ausfluege', 'Anz der Fledermause', 'Luftfeuchtigkeit', 'Temperature']

            for col_index, header in enumerate( headers, start = 1):
                cell = ws.cell(row= 5, column=col_index +2, value= header)
                cell.alignment = Alignment(horizontal= "center", vertical="center")
                cell.fill = header_color
                cell.border = thin_border
                if header =="Datum": 
                    ws.column_dimensions[cell.column_letter].width = len(header) + 10
                else: 
                    ws.column_dimensions[cell.column_letter].width = len(header) + 5
            
            # Write data starting from row 6
            start_row = 6
            start_col = 3
            for row_index, line in enumerate(data_fromTo, start=start_row):
                cell_0 = ws.cell(row=row_index, column=start_col, value=line[0].split(" ")[0])
                cell_1 = ws.cell(row=row_index, column=start_col + 1, value=line[0].split(" ")[1])
                cell_2 = ws.cell(row=row_index, column=start_col + 2, value=int(line[1].strip().replace("->", "")))
                cell_3 = ws.cell(row=row_index, column=start_col + 3, value=int(line[2].strip().replace("<-", "")))
                cell_4 = ws.cell(row=row_index, column=start_col + 4, value=int(line[3].strip().replace("$", "")))
                cell_5 = ws.cell(row=row_index, column=start_col + 5, value=float(line[4].strip().replace("%", "").replace(",", ".")))
                cell_6 = ws.cell(row=row_index, column=start_col + 6, value=float(line[5].strip().replace("C", "").replace(",", ".")))
                for cell in [cell_0, cell_1, cell_2, cell_3, cell_4, cell_5, cell_6]:
                    cell.fill = cell_color
                    cell.border = thin_border


            wb.save(file_path)
            QMessageBox.information(dialog, "Erfolg", f"Daten erfolgreich in '{file_path}' gespeichert")
            dialog.accept()
        except Exception as e:
            QMessageBox.critical(dialog, "Fehler", f"Fehler beim Speichern der Datei: {e}")
            dialog.reject()



